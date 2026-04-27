from __future__ import annotations

import datetime as dt
from dataclasses import dataclass
from typing import Any, Dict, Tuple

import pandas as pd
from openpyxl import load_workbook

_CACHE: Dict[Tuple[str, str, str, str], pd.Series] = {}
_BOND_CACHE: Dict[Tuple[str, str], "BondSeries"] = {}

_CHUNK = 40
_MAX_ROWS = 1200
_MAX_COLS_SCAN = 250

def open_book(path: str):
    # data_only=True => cached calculated values
    return load_workbook(path, data_only=True, read_only=True, keep_vba=True)

def clear_cache():
    _CACHE.clear()
    _BOND_CACHE.clear()

def _to_date(v: Any) -> dt.date | None:
    ts = pd.to_datetime(v, errors="coerce")
    if pd.isna(ts):
        return None
    return ts.date()

def _extract_title(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # allow "Title=...."
    if "Title=" in s:
        # split by semicolon and find Title=
        parts = [p.strip() for p in s.split(";")]
        for p in parts:
            if p.startswith("Title="):
                return p[len("Title="):].strip() or None
    return s

def _match_block(cell_value: Any, block_name: str) -> bool:
    title = _extract_title(cell_value)
    if title is None:
        return False
    return title == block_name

def _find_block_range(ws, block: str) -> tuple[int, int]:
    row1 = [ws.cell(1, c).value for c in range(1, _MAX_COLS_SCAN + 1)]
    starts = []
    for i, v in enumerate(row1, start=1):
        if _match_block(v, block):
            starts.append(i)
    if not starts:
        # 진단용: row1의 non-empty 타이틀 일부 보여주기
        samples = [(i, str(v)) for i, v in enumerate(row1, start=1) if v not in (None, "")]
        head = samples[:20]
        raise ValueError(f"[{ws.title}] block not found: '{block}'. row1 samples={head}")

    start = starts[0]
    # end = next non-empty title -1
    end = _MAX_COLS_SCAN
    for c in range(start + 1, _MAX_COLS_SCAN + 1):
        if row1[c - 1] not in (None, ""):
            end = c - 1
            break
    return start, end

def _build_header_map(ws, start_col: int, end_col: int) -> dict[str, int]:
    hdr = {}
    for c in range(start_col, end_col + 1):
        v = ws.cell(2, c).value
        if v is None:
            continue
        key = str(v).strip()
        if not key:
            continue
        hdr[key] = c
    return hdr

def _read_series(ws, date_col: int, value_col: int, stop_date: dt.date) -> pd.Series:
    idx: list[dt.date] = []
    vals: list[float] = []

    r = 4
    while r <= _MAX_ROWS:
        # chunk read
        for rr in range(r, min(r + _CHUNK, _MAX_ROWS + 1)):
            d = _to_date(ws.cell(rr, date_col).value)
            if d is None:
                continue
            if d < stop_date:
                return pd.Series(vals, index=pd.Index(idx)).sort_index()

            v = pd.to_numeric(ws.cell(rr, value_col).value, errors="coerce")
            if pd.isna(v):
                continue
            idx.append(d)
            vals.append(float(v))

        r += _CHUNK

    if not idx:
        raise ValueError(f"[{ws.title}] no valid rows for date_col={date_col}, value_col={value_col}")
    return pd.Series(vals, index=pd.Index(idx)).sort_index()

class BondSeries:
    def __init__(self, today: pd.Series, prev: pd.Series):
        self.today = today
        self.prev = prev

    def _nearest(self, s: pd.Series, d: dt.date) -> float:
        c = s[s.index <= d]
        return float(c.iloc[-1]) if not c.empty else float("nan")

    def value(self, d: dt.date) -> float:
        return self._nearest(self.today, d)

    def value_prev(self, d: dt.date) -> float:
        # If prev has same-date value, use it; else fallback to prior day in today series
        if d in self.prev.index:
            v = self.prev.loc[d]
            if not pd.isna(v):
                return float(v)
        c = self.today[self.today.index < d]
        return float(c.iloc[-1]) if not c.empty else float("nan")

def load_series(book, sheet: str, block: str, value_col_name: str, asof: dt.date) -> pd.Series:
    key = (sheet, block, value_col_name, asof.isoformat())
    if key in _CACHE:
        return _CACHE[key]

    if sheet not in book.sheetnames:
        raise ValueError(f"missing sheet: {sheet}")
    ws = book[sheet]

    start, end = _find_block_range(ws, block)
    hdr = _build_header_map(ws, start, end)

    if "일자" not in hdr:
        raise ValueError(f"[{sheet}/{block}] missing header '일자'. headers={list(hdr.keys())[:20]}")
    if value_col_name not in hdr:
        raise ValueError(f"[{sheet}/{block}] missing header '{value_col_name}'. headers={list(hdr.keys())[:20]}")

    stop = dt.date(asof.year, 1, 1)
    s = _read_series(ws, hdr["일자"], hdr[value_col_name], stop)
    _CACHE[key] = s
    return s

def load_bond_series(book, sheet: str, block: str, asof: dt.date) -> BondSeries:
    key = (sheet, block)
    if key in _BOND_CACHE:
        return _BOND_CACHE[key]

    if sheet not in book.sheetnames:
        raise ValueError(f"missing sheet: {sheet}")
    ws = book[sheet]
    start, end = _find_block_range(ws, block)
    hdr = _build_header_map(ws, start, end)

    need = ["일자", "민평3사 수익률(산출일) 당일", "민평3사 수익률(산출일)"]
    for n in need:
        if n not in hdr:
            raise ValueError(f"[{sheet}/{block}] missing header '{n}'. headers={list(hdr.keys())[:30]}")

    stop = dt.date(asof.year, 1, 1)
    today = _read_series(ws, hdr["일자"], hdr["민평3사 수익률(산출일) 당일"], stop)
    prev  = _read_series(ws, hdr["일자"], hdr["민평3사 수익률(산출일)"], stop)
    bs = BondSeries(today=today, prev=prev)
    _BOND_CACHE[key] = bs
    return bs